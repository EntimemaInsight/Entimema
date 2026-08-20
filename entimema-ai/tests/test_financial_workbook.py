from datetime import UTC, datetime
from io import BytesIO

from openpyxl import load_workbook

from domain.financial_planning import (
    BusinessStructure,
    CapexSpecification,
    CashFlowSpecification,
    DriverSpecification,
    EvidenceLineage,
    FinancialModelSpecification,
    ModelModule,
    ModelValue,
    PersonnelSpecification,
    PlanningHorizon,
    PlanningScope,
    PlanningType,
    ReportingFrequency,
    ScenarioAssumption,
    ScenarioDefinition,
    WorkingCapitalSpecification,
)
from workbooks.financial import FinancialWorkbookBuilder, WorkbookBuildStatus

NOW = datetime(2027, 1, 2, 3, 4, tzinfo=UTC)


def specification(**updates):
    periods = [f"2027-{month:02d}" for month in range(1, 4)]
    revenue = DriverSpecification(
        driver_id="revenue-model",
        name="Contracts × ACV",
        category="REVENUE",
        formula="contracts * contract_value",
        input_driver_ids=["contracts", "contract_value"],
        output_metric="revenue",
        unit="EUR",
        evidence_ids=["ev-1"],
    )
    cost = DriverSpecification(
        driver_id="cost-model",
        name="Volume × unit cost",
        category="VARIABLE_COST",
        formula="delivery_volume * unit_cost",
        input_driver_ids=["delivery_volume", "unit_cost"],
        output_metric="delivery_cost",
        unit="EUR",
    )
    assumptions = [
        ScenarioAssumption(
            assumption_id=f"{scenario}-{driver}",
            driver_id=driver,
            value=value,
            unit=unit,
            effective_period="2027-01",
            scenario=scenario,
            source="admitted",
            status="ADMITTED",
        )
        for scenario, multiplier in (("Base", 1), ("Upside", 1.1), ("Downside", 0.9))
        for driver, value, unit in (
            ("contracts", 10 * multiplier, "count"),
            ("contract_value", 1000, "EUR"),
            ("delivery_volume", 10 * multiplier, "units"),
            ("unit_cost", 100, "EUR"),
        )
    ]
    values = dict(
        model_id="fm-case-v2",
        case_id="case",
        case_version=2,
        planning_type=PlanningType.ANNUAL_BUDGET,
        scope=PlanningScope(
            horizon=PlanningHorizon(
                forecast_start="2027-01", forecast_end="2027-03", budget_periods=periods
            ),
            currency="EUR",
            frequency=ReportingFrequency.MONTHLY,
            organizational_scope="ServiceCo",
            output_requirements=["P&L", "Cash Flow"],
            scenario_requirements=["Base", "Upside", "Downside"],
        ),
        business_structure=BusinessStructure.SERVICE,
        business_dimensions=["contract"],
        modules=list(ModelModule),
        revenue_drivers=[revenue],
        cost_drivers=[cost],
        personnel=PersonnelSpecification(
            enabled=True,
            opening_fte_driver="contracts",
            hires_driver="contracts",
            leavers_driver="contracts",
            average_fte_formula="average",
            salary_driver="contract_value",
            employer_charges_driver="unit_cost",
        ),
        capex=CapexSpecification(
            enabled=True,
            purchase_value_driver="contract_value",
            useful_life_driver="contracts",
            cash_payment_timing_driver="contract_value",
        ),
        working_capital=WorkingCapitalSpecification(
            enabled=True,
            method="specified balance inputs",
            dso_driver="contracts",
            dpo_driver="contracts",
            inventory_driver="delivery_volume",
            opening_balance_evidence_ids=["ev-1"],
        ),
        cash_flow=CashFlowSpecification(
            enabled=True,
            opening_cash_evidence_id="ev-1",
            operating_cash_sources=["P_AND_L"],
            investing_cash_sources=["CAPEX"],
            liquidity_diagnostics=True,
        ),
        scenarios=[
            ScenarioDefinition(
                scenario_id=x.lower(),
                name=x,
                assumption_ids=[a.assumption_id for a in assumptions if a.scenario == x],
            )
            for x in ("Base", "Upside", "Downside")
        ],
        assumptions=assumptions,
        dependencies=[],
        validation_requirements=[],
        source_evidence_ids=["ev-1"],
        unresolved_issues=[],
        model_name="2027 Budget Model",
        analysis_run_id="run-7b",
        opening_balances=[
            ModelValue(
                metric="opening_cash",
                period="2027-01",
                value=100,
                unit="EUR",
                source="validated opening balance",
                evidence_id="ev-1",
            )
        ],
        actual_values=[
            ModelValue(
                metric=metric,
                period=period,
                value=value,
                unit="EUR",
                source="cash schedule",
                evidence_id="ev-1",
            )
            for period, operating in zip(periods, (-50, -100, 20), strict=True)
            for metric, value in (
                ("operating_cash", operating),
                ("investing_cash", 0),
                ("financing_cash", 0),
            )
        ],
        evidence_lineage=[
            EvidenceLineage(
                evidence_id="ev-1", artifact="FY26.xlsx", source_location="P&L!F27", case_version=2
            )
        ],
    )
    values.update(updates)
    return FinancialModelSpecification(**values)


def test_creates_valid_dynamic_workbook_with_formulas_navigation_and_metadata():
    spec = specification()
    result = FinancialWorkbookBuilder().build(spec, generated_at=NOW)
    assert result.status is WorkbookBuildStatus.PRODUCTION_READY
    assert result.filename == "2027_Budget_Model.xlsx"
    wb = load_workbook(BytesIO(result.workbook_bytes), data_only=False)
    assert wb.sheetnames == [
        "00_Navigation",
        *[
            {
                ModelModule.INPUTS: "01_Assumptions",
                ModelModule.ACTUALS: "02_Actuals",
                ModelModule.REVENUE: "03_Revenue",
                ModelModule.PERSONNEL: "04_Personnel",
                ModelModule.OPEX: "05_OPEX",
                ModelModule.CAPEX: "06_CAPEX",
                ModelModule.WORKING_CAPITAL: "07_Working_Capital",
                ModelModule.PL: "08_P&L",
                ModelModule.CASH_FLOW: "09_Cash_Flow",
                ModelModule.SCENARIOS: "10_Scenarios",
                ModelModule.KPIS: "11_KPIs",
                ModelModule.VALIDATION: "12_Validation",
            }[module]
            for module in ModelModule
        ],
        "_Model_Metadata",
    ]
    assert wb["00_Navigation"]["A11"].hyperlink.target == "#'01_Assumptions'!A1"
    assert wb["03_Revenue"]["B7"].value == "=B5 * B6"
    assert wb["05_OPEX"]["B7"].value == "=B5 * B6"
    assert wb["04_Personnel"]["B8"].value == "=B5+B6-B7"
    assert wb["04_Personnel"]["B9"].value == "=AVERAGE(B5,B8)"
    assert wb["06_CAPEX"]["B7"].value == '=IFERROR(B5/B6,"")'
    assert wb["09_Cash_Flow"]["B10"].value == "=B5+B9"
    assert wb["09_Cash_Flow"]["C5"].value == "=B10"
    assert wb["10_Scenarios"]["B4"].value == "Base"
    assert "SUMIFS" in wb["10_Scenarios"]["E6"].value
    metadata = {row[0].value: row[1].value for row in wb["_Model_Metadata"].iter_rows()}
    assert metadata["formula_structure_validated"] is True
    assert metadata["formula_result_recalculated"] is False
    assert metadata["evidence_references"] == "ev-1"
    assert result.findings[0].first_negative_period == "2027-02"
    assert result.findings[0].minimum_cash == -50
    assert all(check.status != "FAIL" for check in result.validation)


def test_module_selection_and_locked_actual_boundary_are_deterministic():
    spec = specification(
        modules=[
            ModelModule.INPUTS,
            ModelModule.REVENUE,
            ModelModule.SCENARIOS,
            ModelModule.VALIDATION,
        ]
    )
    first = FinancialWorkbookBuilder().build(spec, generated_at=NOW)
    second = FinancialWorkbookBuilder().build(spec, generated_at=NOW)
    assert first.workbook_id == second.workbook_id
    wb = load_workbook(BytesIO(first.workbook_bytes), data_only=False)
    assert wb.sheetnames == [
        "00_Navigation",
        "01_Assumptions",
        "03_Revenue",
        "10_Scenarios",
        "12_Validation",
        "_Model_Metadata",
    ]

    horizon = spec.scope.horizon.model_copy(update={"locked_actual_periods": ["2026-12"]})
    actual = ModelValue(
        metric="contracts",
        period="2026-12",
        value=8,
        unit="count",
        source="actuals",
        evidence_id="ev-1",
    )
    rolling = spec.model_copy(
        update={
            "planning_type": PlanningType.ROLLING_FORECAST,
            "scope": spec.scope.model_copy(update={"horizon": horizon}),
            "actual_values": [actual],
        }
    )
    wb = load_workbook(
        BytesIO(FinancialWorkbookBuilder().build(rolling, generated_at=NOW).workbook_bytes)
    )
    assert wb["03_Revenue"]["B3"].value == "ACTUAL"
    assert wb["03_Revenue"]["B5"].value == 8
    assert wb["03_Revenue"]["C5"].data_type == "f"


def test_missing_opening_cash_is_typed_blocker_and_no_workbook_is_emitted():
    spec = specification(opening_balances=[])
    result = FinancialWorkbookBuilder().build(spec, generated_at=NOW)
    assert result.status is WorkbookBuildStatus.WORKBOOK_BLOCKED
    assert result.missing == ["Opening cash balance"]
    assert result.workbook_bytes is None


def test_generated_workbook_registers_as_artifact_with_download_reference():
    class Registrar:
        def register(self, **kwargs):
            assert kwargs["media_type"].endswith("spreadsheetml.sheet")
            assert kwargs["content"].startswith(b"PK")
            return {
                "artifact": {"id": "artifact-1", "storage_reference": "local://case/artifact-1"}
            }

    result = FinancialWorkbookBuilder().build_and_register(
        specification(),
        registrar=Registrar(),
        owner_id="owner",
        tenant_id="tenant",
        generated_at=NOW,
    )
    assert result.artifact["id"] == "artifact-1"
    assert result.download_reference == "local://case/artifact-1"

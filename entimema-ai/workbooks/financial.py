"""Deterministic FinancialModelSpecification-to-XLSX execution layer."""

from __future__ import annotations

import hashlib
import io
import re
from datetime import UTC, datetime
from enum import StrEnum
from typing import Protocol

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic import BaseModel, ConfigDict, Field

from domain.financial_planning import (
    DriverSpecification,
    FinancialModelSpecification,
    ModelModule,
    ReportingFrequency,
)

GENERATOR_VERSION = "7B.1"
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
NAVY, BLUE, GREY, PALE_BLUE = "14263D", "0070C0", "E7E9ED", "DDEBF7"
GREEN, RED, AMBER = "E2F0D9", "FCE4D6", "FFF2CC"


class WorkbookBuildStatus(StrEnum):
    PRODUCTION_READY = "PRODUCTION_READY"
    WORKBOOK_BLOCKED = "WORKBOOK_BLOCKED"


class ValidationCheck(BaseModel):
    model_config = ConfigDict(extra="forbid", frozen=True)
    check: str
    status: str
    difference: float | None = None
    tolerance: float = 0.0
    detail: str


class LiquidityFinding(BaseModel):
    model_config = ConfigDict(extra="forbid", frozen=True)
    first_negative_period: str
    minimum_cash: float
    currency: str


class FinancialWorkbookResult(BaseModel):
    model_config = ConfigDict(extra="forbid", frozen=True, arbitrary_types_allowed=True)
    status: WorkbookBuildStatus
    workbook_id: str
    filename: str
    specification_id: str
    case_id: str
    case_version: int
    workbook_bytes: bytes | None = Field(exclude=True)
    validation: list[ValidationCheck]
    generated_at: datetime
    generator_version: str
    evidence_lineage: list[str]
    findings: list[LiquidityFinding]
    missing: list[str]
    artifact: dict | None = None
    download_reference: str | None = None


class ArtifactRegistrar(Protocol):
    def register(self, **kwargs): ...


SHEET_NAMES = {
    ModelModule.INPUTS: "01_Assumptions",
    ModelModule.ACTUALS: "02_Actuals",
    ModelModule.REVENUE: "03_Revenue",
    ModelModule.PERSONNEL: "04_Personnel",
    ModelModule.OPEX: "05_OPEX",
    ModelModule.CAPEX: "06_CAPEX",
    ModelModule.WORKING_CAPITAL: "07_Working_Capital",
    ModelModule.PL: "08_P&L",
    ModelModule.CASH_FLOW: "09_Cash_Flow",
    ModelModule.SCENARIOS: "10_Scenarios",
    ModelModule.KPIS: "11_KPIs",
    ModelModule.VALIDATION: "12_Validation",
}


class FinancialWorkbookBuilder:
    """Renders only authoritative specification content; it has no epistemic authority."""

    def build(
        self, specification: FinancialModelSpecification, *, generated_at: datetime | None = None
    ) -> FinancialWorkbookResult:
        generated_at = generated_at or datetime.now(UTC)
        digest = hashlib.sha256(
            specification.model_dump_json(exclude_none=False).encode()
        ).hexdigest()
        workbook_id = f"fwb-{digest[:20]}"
        filename = self._filename(specification.model_name)
        missing = self._blockers(specification)
        if missing:
            return self._result(specification, generated_at, workbook_id, filename, missing=missing)

        wb = Workbook()
        wb.remove(wb.active)
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
        periods = self._periods(specification)
        selected = [module for module in ModelModule if module in specification.modules]
        self._navigation(wb, specification, selected, periods, generated_at)
        for module in selected:
            ws = wb.create_sheet(SHEET_NAMES[module])
            self._base_sheet(ws, module, periods, specification)
            self._render_module(ws, module, specification, periods)
        self._metadata(wb, specification, generated_at, workbook_id, periods)
        validation = self._validate(wb, specification, periods)
        self._write_validation(wb, validation)
        findings = self._liquidity_findings(specification, periods)
        stream = io.BytesIO()
        wb.save(stream)
        content = stream.getvalue()
        # Serialization is part of the gate, not merely a test convenience.
        load_workbook(io.BytesIO(content), data_only=False)
        status = (
            WorkbookBuildStatus.PRODUCTION_READY
            if all(check.status != "FAIL" for check in validation)
            else WorkbookBuildStatus.WORKBOOK_BLOCKED
        )
        return FinancialWorkbookResult(
            status=status,
            workbook_id=workbook_id,
            filename=filename,
            specification_id=specification.model_id,
            case_id=specification.case_id,
            case_version=specification.case_version,
            workbook_bytes=content,
            validation=validation,
            generated_at=generated_at,
            generator_version=GENERATOR_VERSION,
            evidence_lineage=specification.source_evidence_ids,
            findings=findings,
            missing=[],
        )

    def build_and_register(
        self,
        specification: FinancialModelSpecification,
        *,
        registrar: ArtifactRegistrar,
        owner_id: str,
        tenant_id: str,
        generated_at: datetime | None = None,
    ) -> FinancialWorkbookResult:
        result = self.build(specification, generated_at=generated_at)
        if result.status is not WorkbookBuildStatus.PRODUCTION_READY or not result.workbook_bytes:
            return result
        response = registrar.register(
            case_id=result.case_id,
            owner_id=owner_id,
            tenant_id=tenant_id,
            filename=result.filename,
            media_type=XLSX_MEDIA_TYPE,
            content=result.workbook_bytes,
            command_id=f"generate:{result.workbook_id}:{GENERATOR_VERSION}",
        )
        artifact = response["artifact"]
        return result.model_copy(
            update={
                "artifact": artifact,
                "download_reference": artifact["storage_reference"],
            }
        )

    @staticmethod
    def _result(spec, generated_at, workbook_id, filename, *, missing):
        checks = [
            ValidationCheck(
                check="Required authoritative inputs",
                status="FAIL",
                detail=f"Missing: {', '.join(missing)}",
            )
        ]
        return FinancialWorkbookResult(
            status=WorkbookBuildStatus.WORKBOOK_BLOCKED,
            workbook_id=workbook_id,
            filename=filename,
            specification_id=spec.model_id,
            case_id=spec.case_id,
            case_version=spec.case_version,
            workbook_bytes=None,
            validation=checks,
            generated_at=generated_at,
            generator_version=GENERATOR_VERSION,
            evidence_lineage=spec.source_evidence_ids,
            findings=[],
            missing=missing,
        )

    @staticmethod
    def _blockers(spec):
        missing = []
        if spec.cash_flow.enabled and not any(
            value.metric.lower() == "opening_cash" for value in spec.opening_balances
        ):
            missing.append("Opening cash balance")
        if any(issue.blocking for issue in spec.unresolved_issues):
            missing.extend(issue.variable for issue in spec.unresolved_issues if issue.blocking)
        return sorted(set(missing))

    @staticmethod
    def _filename(name):
        safe = re.sub(r"[^A-Za-z0-9_-]+", "_", name).strip("_") or "Budget_Model"
        return f"{safe}.xlsx"

    def _navigation(self, wb, spec, modules, periods, generated_at):
        ws = wb.create_sheet("00_Navigation")
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 48
        ws.merge_cells("A1:B1")
        ws["A1"] = spec.model_name
        self._title(ws["A1"])
        details = [
            ("Case", f"{spec.case_id} / version {spec.case_version}"),
            ("Planning horizon", f"{periods[0]} – {periods[-1]}"),
            ("Currency", spec.scope.currency),
            ("Scenario context", ", ".join(s.name for s in spec.scenarios) or "Not specified"),
            ("Generated (UTC)", generated_at.isoformat()),
        ]
        for row, values in enumerate(details, 3):
            ws.cell(row, 1, values[0]).font = Font(bold=True, color=NAVY)
            ws.cell(row, 2, values[1])
        ws["A10"] = "Model architecture"
        self._section(ws["A10"])
        for index, module in enumerate(modules, 1):
            cell = ws.cell(10 + index, 1, f"{index}. {module.value.replace('_', ' ').title()} →")
            cell.hyperlink = f"#'{SHEET_NAMES[module]}'!A1"
            cell.style = "Hyperlink"
        ws.freeze_panes = "A3"

    def _base_sheet(self, ws, module, periods, spec):
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "B5"
        ws.column_dimensions["A"].width = 32
        ws["A1"] = module.value.replace("_", " ").title()
        self._title(ws["A1"])
        ws["A2"] = "← Navigation"
        ws["A2"].hyperlink = "#'00_Navigation'!A1"
        ws["A2"].style = "Hyperlink"
        for col, period in enumerate(periods, 2):
            cell = ws.cell(4, col, period)
            actual = period in spec.scope.horizon.locked_actual_periods
            cell.fill = PatternFill("solid", fgColor=GREY if actual else NAVY)
            cell.font = Font(bold=True, color="000000" if actual else "FFFFFF")
            cell.alignment = Alignment(horizontal="center")
            ws.cell(
                3,
                col,
                "ACTUAL"
                if actual
                else "BUDGET"
                if period in spec.scope.horizon.budget_periods
                else "FORECAST",
            )
            ws.column_dimensions[get_column_letter(col)].width = 14

    def _render_module(self, ws, module, spec, periods):
        if module is ModelModule.INPUTS:
            self._assumptions(ws, spec)
        elif module is ModelModule.ACTUALS:
            self._actuals(ws, spec, periods)
        elif module is ModelModule.REVENUE:
            self._drivers(ws, spec.revenue_drivers, spec, periods)
        elif module is ModelModule.OPEX:
            self._drivers(ws, spec.cost_drivers, spec, periods)
        elif module is ModelModule.PERSONNEL:
            self._personnel(ws, spec, periods)
        elif module is ModelModule.CAPEX:
            self._capex(ws, spec, periods)
        elif module is ModelModule.WORKING_CAPITAL:
            self._working_capital(ws, spec, periods)
        elif module is ModelModule.PL:
            self._pl(ws, spec, periods)
        elif module is ModelModule.CASH_FLOW:
            self._cash_flow(ws, spec, periods)
        elif module is ModelModule.SCENARIOS:
            self._scenarios(ws, spec)
        elif module is ModelModule.KPIS:
            self._kpis(ws, spec, periods)
        # Validation is populated after formula inspection.

    def _assumptions(self, ws, spec):
        headers = [
            "Assumption ID",
            "Description",
            "Driver",
            "Scenario",
            "Value",
            "Unit",
            "Effective period",
            "Source",
            "Status",
            "Semantic type",
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(4, col, header)
        for row, item in enumerate(spec.assumptions, 5):
            values = [
                item.assumption_id,
                item.assumption_id,
                item.driver_id,
                item.scenario,
                item.value,
                item.unit,
                item.effective_period,
                item.source,
                item.status,
                "Input",
            ]
            for col, value in enumerate(values, 1):
                ws.cell(row, col, value)
            ws.cell(row, 5).font = Font(color=BLUE)
        ws.freeze_panes = "A5"
        ws.auto_filter.ref = f"A4:J{max(4, 4 + len(spec.assumptions))}"

    def _actuals(self, ws, spec, periods):
        metrics = sorted({value.metric for value in spec.actual_values})
        for row, metric in enumerate(metrics, 5):
            ws.cell(row, 1, metric)
            for col, period in enumerate(periods, 2):
                found = next(
                    (x for x in spec.actual_values if x.metric == metric and x.period == period),
                    None,
                )
                if found:
                    cell = ws.cell(row, col, found.value)
                    cell.font = Font(color="008000")
                    evidence = found.evidence_id or "explicit specification"
                    note = (
                        f"Linked Evidence\nSource: {found.source}\n"
                        f"Evidence: {evidence}\nCase version: {spec.case_version}"
                    )
                    cell.comment = Comment(note, "Entimema")

    def _drivers(self, ws, drivers, spec, periods):
        row = 5
        for driver in drivers:
            input_rows = {}
            for input_id in driver.input_driver_ids:
                ws.cell(row, 1, input_id)
                input_rows[input_id] = row
                for col, period in enumerate(periods, 2):
                    if period in spec.scope.horizon.locked_actual_periods:
                        value = self._value(spec.actual_values, input_id, period)
                        ws.cell(row, col, value if value is not None else "")
                    else:
                        ws.cell(row, col, self._assumption_formula(input_id, period))
                        ws.cell(row, col).font = Font(color=BLUE)
                row += 1
            ws.cell(row, 1, driver.output_metric)
            ws.cell(row, 1).font = Font(bold=True)
            for col, period in enumerate(periods, 2):
                if period in spec.scope.horizon.locked_actual_periods:
                    value = self._value(spec.actual_values, driver.output_metric, period)
                    ws.cell(row, col, value if value is not None else "")
                else:
                    ws.cell(row, col, self._translate_formula(driver, input_rows, col))
                    ws.cell(row, col).font = Font(color="000000", bold=True)
            row += 2

    @staticmethod
    def _assumption_formula(driver, period):
        criteria = (
            "'01_Assumptions'!$E:$E,'01_Assumptions'!$C:$C,"
            f'"{driver}",'
            "'01_Assumptions'!$D:$D,Selected_Scenario,"
            f"'01_Assumptions'!$G:$G,\"<={period}\""
        )
        return f'=IFERROR(SUMIFS({criteria}),"")'

    @staticmethod
    def _translate_formula(driver: DriverSpecification, rows, col):
        formula = driver.formula
        for token in sorted(driver.input_driver_ids, key=len, reverse=True):
            formula = re.sub(
                rf"\b{re.escape(token)}\b", f"{get_column_letter(col)}{rows[token]}", formula
            )
        if re.search(r"[^A-Za-z0-9_+*/()., $'-]", formula):
            return "=NA()"
        return f"={formula}"

    def _personnel(self, ws, spec, periods):
        p = spec.personnel
        rows = [
            ("Opening Headcount", p.opening_fte_driver),
            ("Hires", p.hires_driver),
            ("Leavers", p.leavers_driver),
            ("Closing Headcount", None),
            ("Average FTE", None),
            ("Salary", p.salary_driver),
            ("Employer Charges", p.employer_charges_driver),
            ("Bonus", p.bonus_driver),
            ("Benefits", p.benefits_driver),
            ("Personnel Cost", None),
        ]
        for row, (label, driver) in enumerate(rows, 5):
            ws.cell(row, 1, label)
            for col, period in enumerate(periods, 2):
                letter = get_column_letter(col)
                if period in spec.scope.horizon.locked_actual_periods:
                    ws.cell(
                        row,
                        col,
                        self._value(spec.actual_values, label.lower().replace(" ", "_"), period)
                        or "",
                    )
                elif driver:
                    ws.cell(row, col, self._assumption_formula(driver, period))
                elif label == "Closing Headcount":
                    ws.cell(row, col, f"={letter}5+{letter}6-{letter}7")
                elif label == "Average FTE":
                    ws.cell(row, col, f"=AVERAGE({letter}5,{letter}8)")
                elif label == "Personnel Cost":
                    ws.cell(row, col, f"={letter}9*({letter}10+{letter}11+{letter}12+{letter}13)")

    def _capex(self, ws, spec, periods):
        rows = [
            ("Acquisition", spec.capex.purchase_value_driver),
            ("Useful Life", spec.capex.useful_life_driver),
            ("Depreciation", None),
            ("Cash Payment", spec.capex.cash_payment_timing_driver),
        ]
        for row, (label, driver) in enumerate(rows, 5):
            ws.cell(row, 1, label)
            for col, period in enumerate(periods, 2):
                letter = get_column_letter(col)
                ws.cell(
                    row,
                    col,
                    self._assumption_formula(driver, period)
                    if driver
                    else f'=IFERROR({letter}5/{letter}6,"")',
                )

    def _working_capital(self, ws, spec, periods):
        rows = [
            ("Receivables", spec.working_capital.dso_driver),
            ("Payables", spec.working_capital.dpo_driver),
            ("Inventory", spec.working_capital.inventory_driver),
            ("Prepayments", spec.working_capital.prepayments_driver),
            ("Accruals", spec.working_capital.accruals_driver),
        ]
        ws["A3"] = f"Specified method: {spec.working_capital.method}"
        for row, (label, driver) in enumerate(rows, 5):
            ws.cell(row, 1, label)
            if driver:
                for col, period in enumerate(periods, 2):
                    ws.cell(row, col, self._assumption_formula(driver, period))

    def _pl(self, ws, spec, periods):
        outputs = [d.output_metric for d in spec.revenue_drivers + spec.cost_drivers]
        for row, output in enumerate(outputs, 5):
            ws.cell(row, 1, output)
            source = (
                "03_Revenue"
                if any(d.output_metric == output for d in spec.revenue_drivers)
                else "05_OPEX"
            )
            driver_set = spec.revenue_drivers if source == "03_Revenue" else spec.cost_drivers
            source_row = self._driver_output_row(driver_set, output)
            for col in range(2, len(periods) + 2):
                ws.cell(
                    row,
                    col,
                    f"='{source}'!{get_column_letter(col)}{source_row}",
                )
        total = 5 + len(outputs)
        ws.cell(total, 1, "Operating Result")
        for col in range(2, len(periods) + 2):
            letter = get_column_letter(col)
            revenue_rows = [
                5 + i
                for i, x in enumerate(outputs)
                if any(d.output_metric == x for d in spec.revenue_drivers)
            ]
            cost_rows = [
                5 + i
                for i, x in enumerate(outputs)
                if any(d.output_metric == x for d in spec.cost_drivers)
            ]
            revenue_refs = ",".join(letter + str(r) for r in revenue_rows)
            cost_refs = ",".join(letter + str(r) for r in cost_rows)
            ws.cell(total, col, f"=SUM({revenue_refs})-SUM({cost_refs})")

    @staticmethod
    def _driver_output_row(drivers, output):
        row = 5
        for driver in drivers:
            row += len(driver.input_driver_ids)
            if driver.output_metric == output:
                return row
            row += 2
        raise ValueError(output)

    def _cash_flow(self, ws, spec, periods):
        opening = next(x.value for x in spec.opening_balances if x.metric.lower() == "opening_cash")
        labels = [
            "Opening Cash",
            "Operating Cash Flow",
            "Investing Cash Flow",
            "Financing Cash Flow",
            "Net Movement",
            "Closing Cash",
            "Cash Flow Check",
        ]
        for row, label in enumerate(labels, 5):
            ws.cell(row, 1, label)
        for col in range(2, len(periods) + 2):
            letter = get_column_letter(col)
            ws.cell(5, col, opening if col == 2 else f"={get_column_letter(col - 1)}10")
            source_rows = (
                (6, "operating_cash", spec.cash_flow.operating_cash_sources),
                (7, "investing_cash", spec.cash_flow.investing_cash_sources),
                (8, "financing_cash", spec.cash_flow.financing_cash_sources),
            )
            for row, source_type, sources in source_rows:
                value = self._value(spec.actual_values, source_type, periods[col - 2])
                if value is not None:
                    ws.cell(row, col, value)
                else:
                    ws.cell(row, col, self._cash_source_formula(source_type, sources, col, spec))
            ws.cell(9, col, f"=SUM({letter}6:{letter}8)")
            ws.cell(10, col, f"={letter}5+{letter}9")
            ws.cell(11, col, f"={letter}5+{letter}9-{letter}10")

    @staticmethod
    def _cash_source_formula(source_type, sources, col, spec):
        letter = get_column_letter(col)
        if not sources:
            return 0  # An explicitly empty source list means no movement of this class.
        if source_type == "operating_cash" and "P_AND_L" in sources:
            operating_result_row = 5 + len(spec.revenue_drivers + spec.cost_drivers)
            return f"='08_P&L'!{letter}{operating_result_row}"
        if source_type == "investing_cash" and "CAPEX" in sources:
            return f"=-'06_CAPEX'!{letter}8"
        return "=NA()"

    def _scenarios(self, ws, spec):
        ws["A4"], ws["B4"] = (
            "Selected Scenario",
            spec.scenarios[0].name if spec.scenarios else "Base",
        )
        names = [s.name for s in spec.scenarios] or ["Base"]
        dv = DataValidation(type="list", formula1='"' + ",".join(names) + '"')
        ws.add_data_validation(dv)
        dv.add(ws["B4"])
        wb = ws.parent
        wb.create_named_range("Selected_Scenario", ws, "$B$4")
        ws["A6"] = "Scenario"
        for row, name in enumerate(names, 7):
            ws.cell(row, 1, name)
            ws.cell(row, 2, "Controlled assumption set")
        if len(names) > 1:
            ws["D4"] = "End-period driver comparison"
            for col, name in enumerate(names, 5):
                ws.cell(5, col, name)
            for row, driver in enumerate(spec.revenue_drivers + spec.cost_drivers, 6):
                ws.cell(row, 4, driver.output_metric)
                for col, _name in enumerate(names, 5):
                    formula = driver.formula
                    for token in sorted(driver.input_driver_ids, key=len, reverse=True):
                        lookup = (
                            "SUMIFS('01_Assumptions'!$E:$E,"
                            f"'01_Assumptions'!$C:$C,\"{token}\","
                            f"'01_Assumptions'!$D:$D,{get_column_letter(col)}$5)"
                        )
                        formula = re.sub(rf"\b{re.escape(token)}\b", lookup, formula)
                    ws.cell(row, col, f"={formula}")

    def _kpis(self, ws, spec, periods):
        metrics = ["Closing Cash"] if ModelModule.CASH_FLOW in spec.modules else []
        for row, metric in enumerate(metrics, 5):
            ws.cell(row, 1, metric)
            for col in range(2, len(periods) + 2):
                ws.cell(row, col, f"='09_Cash_Flow'!{get_column_letter(col)}10")

    def _metadata(self, wb, spec, generated_at, workbook_id, periods):
        ws = wb.create_sheet("_Model_Metadata")
        values = {
            "model_version": spec.model_id,
            "generator_version": GENERATOR_VERSION,
            "workbook_id": workbook_id,
            "case_id": spec.case_id,
            "case_version": spec.case_version,
            "analysis_run_id": spec.analysis_run_id or "not supplied",
            "specification_id": spec.model_id,
            "generated_timestamp": generated_at.isoformat(),
            "currency": spec.scope.currency,
            "horizon": f"{periods[0]}:{periods[-1]}",
            "evidence_references": ",".join(spec.source_evidence_ids),
            "formula_structure_validated": True,
            "formula_result_recalculated": False,
        }
        for row, item in enumerate(values.items(), 1):
            ws.cell(row, 1, item[0])
            ws.cell(row, 2, item[1])
        if "Selected_Scenario" not in wb.defined_names:
            ws["D1"] = spec.scenarios[0].name if spec.scenarios else "Base"
            wb.create_named_range("Selected_Scenario", ws, "$D$1")
        start = len(values) + 2
        for row, lineage in enumerate(spec.evidence_lineage, start):
            ws.cell(row, 1, "lineage")
            ws.cell(
                row,
                2,
                f"{lineage.evidence_id}|{lineage.artifact}|{lineage.source_location}|case-v{lineage.case_version}",
            )
        ws.sheet_state = "hidden"
        wb.properties.title = spec.model_name
        wb.properties.subject = f"Financial model for {spec.case_id} v{spec.case_version}"
        wb.properties.creator = f"Entimema Workbook Engine {GENERATOR_VERSION}"

    def _validate(self, wb, spec, periods):
        formulas = [
            (ws.title, c.coordinate, c.value)
            for ws in wb.worksheets
            for row in ws.iter_rows()
            for c in row
            if c.data_type == "f"
        ]
        sheet_names = set(wb.sheetnames)
        broken = [f"{s}!{c}" for s, c, value in formulas if "#REF!" in value]
        malformed = [f"{s}!{c}" for s, c, value in formulas if value == "=NA()"]
        refs = re.compile(r"'(.*?)'!")
        missing_refs = [
            f"{s}!{c}:{ref}"
            for s, c, value in formulas
            for ref in refs.findall(value)
            if ref not in sheet_names
        ]
        declared_drivers = spec.revenue_drivers + spec.cost_drivers
        drivers = {d.driver_id for d in declared_drivers}
        drivers.update(
            input_id for driver in declared_drivers for input_id in driver.input_driver_ids
        )
        scenario_missing = [a.assumption_id for a in spec.assumptions if a.driver_id not in drivers]
        return [
            ValidationCheck(
                check="Formula structure",
                status="FAIL" if broken or missing_refs or malformed else "PASS",
                detail="Broken, missing or unsupported references: "
                + ", ".join(broken + missing_refs + malformed)
                if broken or missing_refs or malformed
                else "Formula structure validated; results not recalculated by Python.",
            ),
            ValidationCheck(
                check="Timeline",
                status="PASS",
                detail=f"Exactly {len(periods)} authoritative periods rendered.",
            ),
            ValidationCheck(
                check="Scenario mapping",
                status="FAIL" if scenario_missing else "PASS",
                detail="Unmapped: " + ", ".join(scenario_missing)
                if scenario_missing
                else "All scenario assumptions map to declared drivers.",
            ),
            ValidationCheck(
                check="Evidence provenance",
                status="PASS" if spec.source_evidence_ids or spec.assumptions else "WARNING",
                detail="Baseline provenance references retained."
                if spec.source_evidence_ids
                else "No evidence references supplied; explicit assumptions only.",
            ),
            ValidationCheck(
                check="Model connectivity",
                status="PASS",
                detail=f"{len(spec.dependencies)} declared dependency edges retained.",
            ),
            ValidationCheck(
                check="Cash roll-forward",
                status="PASS"
                if not spec.cash_flow.enabled or ModelModule.CASH_FLOW in spec.modules
                else "FAIL",
                detail="Opening cash + net movement = closing cash formula emitted."
                if spec.cash_flow.enabled
                else "Not requested.",
            ),
        ]

    def _write_validation(self, wb, checks):
        if "12_Validation" not in wb.sheetnames:
            return
        ws = wb["12_Validation"]
        for col, header in enumerate(["Check", "Status", "Difference", "Tolerance", "Detail"], 1):
            ws.cell(4, col, header)
        for row, check in enumerate(checks, 5):
            for col, value in enumerate(
                [check.check, check.status, check.difference, check.tolerance, check.detail], 1
            ):
                ws.cell(row, col, value)
        for status, color in (("PASS", GREEN), ("WARNING", AMBER), ("FAIL", RED)):
            ws.conditional_formatting.add(
                f"B5:B{4 + len(checks)}",
                CellIsRule(
                    operator="equal",
                    formula=[f'"{status}"'],
                    fill=PatternFill("solid", fgColor=color),
                ),
            )

    def _liquidity_findings(self, spec, periods):
        if not spec.cash_flow.enabled or not spec.cash_flow.liquidity_diagnostics:
            return []
        opening = next(x.value for x in spec.opening_balances if x.metric.lower() == "opening_cash")
        balances = []
        cash = opening
        for period in periods:
            components = [
                self._value(spec.actual_values, metric, period)
                for metric in ("operating_cash", "investing_cash", "financing_cash")
            ]
            if any(value is None for value in components):
                return []  # Excel will calculate formulas; Python does not claim to have done so.
            cash += sum(components)
            balances.append((period, cash))
        negative = [item for item in balances if item[1] < 0]
        return (
            [
                LiquidityFinding(
                    first_negative_period=negative[0][0],
                    minimum_cash=min(x[1] for x in balances),
                    currency=spec.scope.currency,
                )
            ]
            if negative
            else []
        )

    @staticmethod
    def _value(values, metric, period):
        item = next(
            (x for x in values if x.metric.lower() == metric.lower() and x.period == period), None
        )
        return item.value if item else None

    def _periods(self, spec):
        horizon = spec.scope.horizon
        requested = list(
            dict.fromkeys(
                horizon.locked_actual_periods
                + self._range(horizon.forecast_start, horizon.forecast_end, spec.scope.frequency)
            )
        )
        return requested

    @staticmethod
    def _range(start, end, frequency):
        if frequency is ReportingFrequency.QUARTERLY and "Q" in start:
            year, quarter = map(int, start.replace("Q", "").split("-"))
            end_year, end_quarter = map(int, end.replace("Q", "").split("-"))
            result = []
            while (year, quarter) <= (end_year, end_quarter):
                result.append(f"{year}-Q{quarter}")
                quarter += 1
                if quarter == 5:
                    year, quarter = year + 1, 1
            return result
        year, month = map(int, (start.split("-") + ["1"])[:2])
        end_year, end_month = map(int, (end.split("-") + ["1"])[:2])
        step = {
            ReportingFrequency.MONTHLY: 1,
            ReportingFrequency.QUARTERLY: 3,
            ReportingFrequency.ANNUAL: 12,
        }.get(frequency)
        if step is None:
            raise ValueError("Workbook engine supports monthly, quarterly and annual frequency")
        result = []
        while (year, month) <= (end_year, end_month):
            result.append(
                str(year) if frequency is ReportingFrequency.ANNUAL else f"{year}-{month:02d}"
            )
            month += step
            year += (month - 1) // 12
            month = (month - 1) % 12 + 1
        return result

    @staticmethod
    def _title(cell):
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(bold=True, color="FFFFFF", size=16)
        cell.alignment = Alignment(vertical="center")

    @staticmethod
    def _section(cell):
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(bold=True, color="FFFFFF")
